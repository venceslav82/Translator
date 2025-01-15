from flask import Flask, request, render_template, send_file
import os
from googletrans import Translator
from docx import Document
from openpyxl import load_workbook
import xlrd
from PyPDF2 import PdfReader
import tempfile

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
TRANSLATED_FOLDER = 'translated'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TRANSLATED_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

translator = Translator()

def translate_text(content, src='en', dest='bg'):
    """Превежда даден текст."""
    if not content:  # Проверка за празно съдържание
        return ''
    max_len = 5000  # Лимит за дължина на текста
    translated = []
    for i in range(0, len(content), max_len):
        part = content[i:i + max_len]
        try:
            result = translator.translate(part, src=src, dest=dest)
            translated.append(result.text)
        except Exception as e:
            print(f"Translation error: {e}")
            translated.append(part)  # Ако има грешка, връщаме оригиналния текст
    return ''.join(translated)

def handle_docx(filepath, translated_path):
    """Обработва и превежда .docx файлове."""
    doc = Document(filepath)
    for paragraph in doc.paragraphs:
        paragraph.text = translate_text(paragraph.text)
    doc.save(translated_path)

def handle_xlsx(filepath, translated_path):
    """Обработва и превежда .xlsx файлове."""
    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = translate_text(cell.value)
    wb.save(translated_path)

def handle_xls(filepath, translated_path):
    """Обработва и превежда .xls файлове."""
    book = xlrd.open_workbook(filepath)
    translated_wb = load_workbook(translated_path)
    for sheet_idx in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_idx)
        translated_ws = translated_wb.create_sheet(sheet.name)
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                if isinstance(value, str):
                    value = translate_text(value)
                translated_ws.cell(row=row_idx+1, column=col_idx+1, value=value)
    translated_wb.save(translated_path)

def handle_pdf(filepath, translated_path):
    """Обработва и превежда .pdf файлове."""
    reader = PdfReader(filepath)
    translated_content = []
    for page in reader.pages:
        translated_content.append(translate_text(page.extract_text()))
    with open(translated_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(translated_content))

@app.route('/')
def upload_file():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def handle_upload():
    file = request.files['file']
    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        ext = os.path.splitext(file.filename)[1].lower()
        translated_path = os.path.join(TRANSLATED_FOLDER, f'translated_{file.filename}')
        
        try:
            if ext == '.docx':
                handle_docx(filepath, translated_path)
            elif ext in ['.xls', '.xlsx']:
                handle_xlsx(filepath, translated_path) if ext == '.xlsx' else handle_xls(filepath, translated_path)
            elif ext == '.pdf':
                handle_pdf(filepath, translated_path)
            else:
                return 'Unsupported file type!', 400
        except Exception as e:
            return f'Error processing file: {e}', 500

        return send_file(translated_path, as_attachment=True)

    return 'No file uploaded!', 400

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
